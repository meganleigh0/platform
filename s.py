import openpyxl
import pandas as pd
import os

# Function to load and parse data from an Excel workbook for each sheet (each day)
def load_workbook_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    all_data = []

    # Loop through each sheet (each day)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Extract data for Plant 1, Plant 3, and TNA
        sheet_data = extract_plant_data(ws, sheet_name)
        if sheet_data is not None:
            all_data.append(sheet_data)

    # Combine all data for the workbook into one DataFrame
    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()

# Function to extract data from Plant 1, Plant 3, and TNA sections in a sheet
def extract_plant_data(ws, sheet_name):
    sections = ["P1 Turrets", "P1 Hall", "P3 Turrets", "P3 Hall", "TNA"]

    data_list = []
    # Example station names
    plant1_turret_stations = ["Machining", "Armor Install", "Appurtenance", "Paint"]
    plant1_hall_stations = ["Armor", "Structure", "Appurtenance", "Paint"]
    plant3_stations = ["Station 0", "Station 1", "Stations 2-3-4", "Stations 5-6", "Stations 7-8"]
    tna_stations = ["Test and Accept", "Final Paint", "Prep and Chips"]

    # Parse each plant's section based on key phrases and extract summary tables and vehicles
    plant1_data = parse_plant_section(ws, "P1 Turrets", plant1_turret_stations, sheet_name)
    plant1_data += parse_plant_section(ws, "P1 Hall", plant1_hall_stations, sheet_name)

    plant3_data = parse_plant_section(ws, "P3 Turrets", plant3_stations, sheet_name)
    plant3_data += parse_plant_section(ws, "P3 Hall", plant3_stations, sheet_name)

    tna_data = parse_plant_section(ws, "TNA", tna_stations, sheet_name)

    # Combine all plant data
    return pd.concat(plant1_data + plant3_data + tna_data, ignore_index=True) if plant1_data or plant3_data or tna_data else None

# Function to parse individual sections of each plant
def parse_plant_section(ws, section_name, stations, sheet_name):
    section_data = []
    for station in stations:
        start_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if any(station in str(cell.value) for cell in row):
                start_row = row[0].row
                break

        if start_row:
            # Extract summary table and vehicles data for the station
            summary_data, vehicle_data = extract_summary_and_vehicle_data(ws, start_row)
            # Add Station, Section, and Date information
            summary_data['Station'] = station
            summary_data['Section'] = section_name
            summary_data['Date'] = sheet_name
            vehicle_data['Station'] = station
            vehicle_data['Section'] = section_name
            vehicle_data['Date'] = sheet_name

            # Append the summary and vehicle data
            section_data.append(summary_data)
            section_data.append(vehicle_data)

    return section_data

# Function to extract the summary table and vehicle data for a station
def extract_summary_and_vehicle_data(ws, start_row):
    summary_data = []
    vehicle_data = []
    collecting_summary = True
    current_data = []

    for row in ws.iter_rows(min_row=start_row + 1, max_row=ws.max_row, values_only=True):
        if collecting_summary:
            if any(row):  # If row has content, it's part of the summary table
                current_data.append(row)
            else:
                # If row is empty, we switch to vehicle data
                collecting_summary = False
                summary_data = pd.DataFrame(current_data, columns=['Contract', 'MRP', 'Actual', 'Delta', 'Flow'])
                current_data = []
        else:
            # Collect vehicle data (will stop on encountering "M days" or "Planned WIP")
            if "M days" in str(row) or "Planned WIP" in str(row):
                break  # End of the vehicle list
            current_data.append(row)

    vehicle_data = pd.DataFrame(current_data, columns=['Vehicles']) if current_data else pd.DataFrame()
    return summary_data, vehicle_data

# Function to load all workbooks in the "DailyStatus" folder
def load_all_workbooks(directory_path):
    all_data = []

    # Loop through each workbook (e.g., DailyStatusArchivesAPR2024, DailyStatusArchivesFEB2024)
    for file_name in os.listdir(directory_path):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(directory_path, file_name)
            workbook_data = load_workbook_data(file_path)
            if not workbook_data.empty:
                all_data.append(workbook_data)

    # Combine all data from all workbooks into one DataFrame
    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()

# Example usage
directory = "DailyStatus"  # Folder containing your daily status files
final_data = load_all_workbooks(directory)

# Display the final aggregated data
print(final_data)